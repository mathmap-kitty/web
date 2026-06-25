# -*- coding: utf-8 -*-
"""跨單元整合題 · 脈絡地圖。

資料來源：參考文件/學測數學考題分析_106-115_更正版.xlsx 的「跨單元清單」分頁。
題目資料（年度/題號/單元配對/核心概念/難易）逐筆取自 Excel；
分組主題與「脈絡」敘述為教學編排。視覺重用 build/assets/style.css。
"""
import os, io, sys
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
sys.path.insert(0, HERE)
from build_html import ANALYTICS  # noqa: E402  共用 GA4 + Clarity 追蹤碼
KATEX = "https://cdnjs.cloudflare.com/ajax/libs/KaTeX/0.16.9"
XLSX = os.path.join(ROOT, "參考文件", "學測數學考題分析_106-115_更正版.xlsx")
OUT = os.path.join(ROOT, "dist", "115學測數學_跨單元整合_脈絡地圖.html")

sys.path.insert(0, os.path.join(ROOT, "content"))
sys.path.insert(0, HERE)
from render import html_rich


def _asset(name):
    return io.open(os.path.join(HERE, "assets", name), encoding="utf-8").read()


# 單元名稱 → 對應單元頁（可點按跳轉）
UNIT_FILE = {
    "數與式": "115學測數學_數與式.html",
    "多項式函數": "115學測數學_多項式函數.html",
    "指數與對數": "115學測數學_指數與對數.html",
    "數列級數": "115學測數學_數列與級數.html",
    "排列組合": "115學測數學_排列組合與機率.html",
    "機率期望值": "115學測數學_排列組合與機率.html",
    "數據分析": "115學測數學_數據分析.html",
    "三角比": "115學測數學_三角.html",
    "三角函數": "115學測數學_三角.html",
    "平面向量": "115學測數學_平面向量.html",
    "空間向量": "115學測數學_空間向量.html",
    "空間的平面與直線": "115學測數學_空間向量.html",
    "矩陣": "115學測數學_矩陣.html",
}

# 主題（脈絡）→ 成員題目 key（"學年度|卷別|題號"），各主題內依年度排序
THEMES = [
    ("① 向量：跨單元之王",
     "向量最會「串門子」——它把長度、角度、面積全化成坐標與內積運算，於是能和三角、多項式、數與式、指對數通通結合。看到向量題，先想「能不能坐標化」。",
     ["107|數A|10", "107|數A|G", "108|數A|G", "111|數B|14",
      "111|數B|4", "113|數B|4", "115|數A|14"]),
    ("② 三角：比與函數的縫合",
     "「三角比」管直角三角形與測量、「三角函數」管廣義角與恆等式。學測常把兩者縫在一起：先用正/餘弦定理算邊角，再用倍角、和角化簡——多項式的對稱性也常來軋一腳。",
     ["109|數A|G", "113|數A|6", "115|數A|17", "109|數A|7"]),
    ("③ 數列：規律與成長的離散版",
     "數列是「離散的函數」。它的規律常藏在二次函數、指數成長（指對數）裡；一旦要求和，又會碰上計數原理——所以數列題天生愛跨界。",
     ["106|數A|A", "107|數A|5", "110|數A|2", "113|數A|8",
      "113|數B|17", "115|數A|3"]),
    ("④ 機率：計數與資料的十字路口",
     "機率的分母來自排列組合的計數，分子常牽涉矩陣（解聯立）、指對數（估量級）或數據（列聯表）。它正站在「計數 × 資料」的交會點。",
     ["109|數A|6", "113|數A|11", "114|數B|4", "114|數B|13", "115|數B|15"]),
    ("⑤ 空間：把平面升一個維度",
     "空間題把平面的畢氏、餘弦定理、向量「升一維」，再加上點到平面距離、外積、二面角——常是壓軸難題，務必先畫圖、建坐標。",
     ["107|數A|H", "111|數A|17", "111|數B|6"]),
]

DIFF_CLASS = {"易": "easy", "中": "mid", "難": "hard"}

PVEC = "115學測數學_平面向量.html"
TRIG = "115學測數學_三角.html"
POLY = "115學測數學_多項式函數.html"
PROB = "115學測數學_排列組合與機率.html"
EXPLOG = "115學測數學_指數與對數.html"
SPACE = "115學測數學_空間向量.html"
MATRIX = "115學測數學_矩陣.html"
SEQ = "115學測數學_數列與級數.html"
NUMEXPR = "115學測數學_數與式.html"
DATA = "115學測數學_數據分析.html"

# 互動題目（完整題幹＋簡答＋解題關鍵＋直達考點）。題幹取自官方原卷、答案對官方答案頁校正。
# key = "學年度|卷別|題號"。目前先完成「向量」脈絡的數A題作為範本。
DETAIL = {
    "107|數A|10": {
        "body": r"已知坐標平面上 \(\triangle ABC\)，其中 \(\overrightarrow{AB}=(-4,3)\)，且 \(\overrightarrow{AC}=\left(\dfrac{2}{5},\dfrac{4}{5}\right)\)。試選出正確的選項。",
        "opts": [r"\(\overline{BC}=5\)", r"\(\triangle ABC\) 是直角三角形",
                 r"\(\triangle ABC\) 的面積為 \(\dfrac{11}{5}\)", r"\(\sin B>\sin C\)", r"\(\cos A>\cos B\)"],
        "ans": r"**(2)(3)**",
        "key": [r"\(\overrightarrow{BC}=\overrightarrow{AC}-\overrightarrow{AB}=\left(\tfrac{22}{5},-\tfrac{11}{5}\right)\)，\(|\overrightarrow{BC}|=\tfrac{11\sqrt5}{5}\neq5\)（(1) 錯）；",
                r"面積 \(=\tfrac12\left|(-4)\cdot\tfrac45-3\cdot\tfrac25\right|=\dfrac{11}{5}\)（(3) 對）；",
                r"\(\overrightarrow{CA}\cdot\overrightarrow{CB}=0\Rightarrow\angle C=90^\circ\)（(2) 對）；由直角 \(C\) 推得 (4)(5) 皆錯。"],
        "links": [(r"平面向量 ▸ 考點1 向量運算", PVEC, "kp1"), (r"三角 ▸ 考點2 正餘弦定理", TRIG, "kp2")],
    },
    "107|數A|G": {
        "body": r"設 \(D\) 為 \(\triangle ABC\) 中 \(\overline{BC}\) 邊上的一點，已知 \(\angle ABC=75^\circ\)、\(\angle ACB=45^\circ\)、\(\angle ADB=60^\circ\)。若 \(\overrightarrow{AD}=s\,\overrightarrow{AB}+t\,\overrightarrow{AC}\)，求 \(s,t\)。",
        "ans": r"**\(s=\dfrac13,\quad t=\dfrac23\)**",
        "key": [r"\(D\) 在 \(\overline{BC}\) 上 \(\Rightarrow\overrightarrow{AD}=(1-k)\overrightarrow{AB}+k\overrightarrow{AC}\)，所以 \(s+t=1\)；",
                r"正弦定理算分點比：\(\dfrac{BD}{DC}=\dfrac{\sin^2 45^\circ}{\sin75^\circ\sin15^\circ}=2\Rightarrow k=\tfrac23\)；",
                r"故 \(\overrightarrow{AD}=\tfrac13\overrightarrow{AB}+\tfrac23\overrightarrow{AC}\)，即 \(s=\tfrac13,\ t=\tfrac23\)。"],
        "links": [(r"三角 ▸ 考點2 正弦定理", TRIG, "kp2"), (r"平面向量 ▸ 考點2 線性組合·分點", PVEC, "kp2")],
    },
    "108|數A|G": {
        "body": r"如圖，\(A,B,C,D\) 為平面上四點。已知 \(\overrightarrow{BC}=\overrightarrow{AB}+\overrightarrow{AD}\)，且 \(\overrightarrow{AC}\)、\(\overrightarrow{BD}\) 兩向量 **等長且互相垂直**，求 \(\tan\angle BAD\)。",
        "fig": r'''<svg viewBox="0 0 232 188" width="232" height="188" xmlns="http://www.w3.org/2000/svg" font-family="system-ui,'Segoe UI',sans-serif">
  <polygon points="34,98 122,30 202,90 110,162" fill="#eef6f4" stroke="#1f6f78" stroke-width="1.6"/>
  <line x1="34" y1="98" x2="202" y2="90" stroke="#c0392b" stroke-width="1.7"/>
  <line x1="122" y1="30" x2="110" y2="162" stroke="#1f3a93" stroke-width="1.7"/>
  <g fill="#2c3e50"><circle cx="34" cy="98" r="2.6"/><circle cx="122" cy="30" r="2.6"/><circle cx="202" cy="90" r="2.6"/><circle cx="110" cy="162" r="2.6"/></g>
  <g font-size="14" font-weight="bold" font-style="italic" fill="#22404a">
    <text x="16" y="103">A</text><text x="118" y="22">D</text><text x="208" y="93">C</text><text x="100" y="180">B</text>
  </g>
  <text x="150" y="70" font-size="10.5" fill="#c0392b" font-style="italic">AC</text>
  <text x="124" y="104" font-size="10.5" fill="#1f3a93" font-style="italic">BD</text>
</svg>''',
        "ans": r"**\(\tan\angle BAD=-3\)**",
        "key": [r"令 \(\overrightarrow{AB}=\vec u,\ \overrightarrow{AD}=\vec v\)：\(\overrightarrow{AC}=2\vec u+\vec v\)、\(\overrightarrow{BD}=\vec v-\vec u\)；",
                r"垂直 \((2\vec u+\vec v)\cdot(\vec v-\vec u)=0\)、等長 \(|2\vec u+\vec v|=|\vec v-\vec u|\)；",
                r"解得 \(\vec u\cdot\vec v=-\tfrac12|\vec u|^2,\ |\vec v|^2=\tfrac52|\vec u|^2\)，故 \(\cos\angle BAD=-\tfrac{1}{\sqrt{10}},\ \tan\angle BAD=-3\)。"],
        "links": [(r"平面向量 ▸ 考點3 內積", PVEC, "kp3"), (r"三角 ▸ 考點1 三角比定義", TRIG, "kp1")],
    },
    "115|數A|14": {
        "body": r"坐標平面上，向量 \((a,b)\) 與直線 \(y=bx-1\) 垂直，則 \(a+b\) 的最大可能值為何？（最簡分數）",
        "ans": r"**\(\dfrac14\)**",
        "key": [r"直線 \(y=bx-1\) 方向向量 \((1,b)\)，垂直 \(\Rightarrow(a,b)\cdot(1,b)=a+b^2=0\Rightarrow a=-b^2\)；",
                r"\(a+b=-b^2+b=-\left(b-\tfrac12\right)^2+\tfrac14\)（配方）；",
                r"當 \(b=\tfrac12\) 取最大值 \(\dfrac14\)。"],
        "links": [(r"平面向量 ▸ 考點3 內積·垂直", PVEC, "kp3"), (r"多項式 ▸ 考點2 二次配方求極值", POLY, "kp2")],
    },
    "109|數A|7": {
        "body": r"坐標平面上，函數圖形 \(y=-\sqrt3\,x^3\) 上有兩點 \(P,Q\) 到原點距離皆為 \(1\)。已知點 \(P\) 坐標為 \((\cos\theta,\sin\theta)\)，試問點 \(Q\) 坐標為何？",
        "opts": [r"\((\cos(-\theta),\sin(-\theta))\)", r"\((-\cos\theta,\sin\theta)\)",
                 r"\((\cos(-\theta),-\sin\theta)\)", r"\((-\cos\theta,\sin(-\theta))\)", r"\((\cos\theta,-\sin\theta)\)"],
        "ans": r"**(4)** \((-\cos\theta,\sin(-\theta))\)",
        "key": [r"\(y=-\sqrt3\,x^3\) 是 **奇函數**，圖形對 **原點對稱**；",
                r"單位圓與此曲線交於兩點 \(P,Q\)，必為原點對稱 \(\Rightarrow Q=-P=(-\cos\theta,-\sin\theta)\)；",
                r"\(-\sin\theta=\sin(-\theta)\)，故 \(Q=(-\cos\theta,\sin(-\theta))\)，選 (4)。"],
        "links": [(r"多項式 ▸ 考點5 三次函數對稱", POLY, "kp5"), (r"三角 ▸ 考點1 廣義角", TRIG, "kp1")],
    },
    "109|數A|6": {
        "body": r"連續投擲一公正骰子兩次，設出現的點數依序為 \(a,b\)。試問發生 \(\log(a^2)+\log b>1\) 的機率為多少？",
        "opts": [r"\(\dfrac13\)", r"\(\dfrac12\)", r"\(\dfrac23\)", r"\(\dfrac34\)", r"\(\dfrac56\)"],
        "ans": r"**(4)** \(\dfrac34\)",
        "key": [r"\(\log(a^2)+\log b=\log(a^2b)>1\Leftrightarrow a^2b>10\)；",
                r"逐一數 \(a^2b>10\)：\(a=1\) 有 \(0\)、\(a=2\) 有 \(4\)、\(a=3\) 有 \(5\)、\(a\ge4\) 各 \(6\) 個，共 \(27\)；",
                r"機率 \(=\dfrac{27}{36}=\dfrac34\)。"],
        "links": [(r"指數對數 ▸ 考點2 對數運算", EXPLOG, "kp2"), (r"排列機率 ▸ 考點3 古典機率", PROB, "kp3")],
    },
    "113|數A|11": {
        "body": r"考慮二元一次方程組 \(\begin{cases}ax+6y=6\\x+by=1\end{cases}\)，係數 \(a,b\) 由投擲一顆公正骰子與一枚均勻硬幣決定：\(a\) 為骰子點數；硬幣正面 \(b=1\)、反面 \(b=2\)。試選出正確的選項。",
        "opts": [r"擲出 \(a=b\) 的機率為 \(\tfrac13\)", r"方程組無解的機率為 \(\tfrac1{12}\)",
                 r"恰有一組解的機率為 \(\tfrac56\)", r"硬幣反面且方程組有解的機率為 \(\tfrac12\)",
                 r"在硬幣反面且有解的條件下，\(a\) 值符合某條件的機率為 \(\tfrac25\)"],
        "ans": r"**(2)(3)**",
        "key": [r"樣本空間 \(n(S)=6\times2=12\)；當係數比 \(\tfrac{a}{1}=\tfrac{6}{b}\) 時兩直線平行或重合；",
                r"無解（平行）只在 \((a,b)=(3,2)\)，機率 \(\tfrac1{12}\)（(2) 對）；重合 \((6,1)\) 也 \(\tfrac1{12}\)；",
                r"恰一解 \(=1-\tfrac1{12}-\tfrac1{12}=\tfrac{10}{12}=\tfrac56\)（(3) 對）。"],
        "links": [(r"矩陣 ▸ 考點4 一次聯立解的判別", MATRIX, "kp4"), (r"排列機率 ▸ 考點4 條件機率", PROB, "kp4")],
    },
    "107|數A|H": {
        "body": r"將一塊邊長 \(\overline{AB}=15\) 公分、\(\overline{BC}=20\) 公分的長方形鐵片 \(ABCD\) 沿對角線 \(\overline{BD}\) 對摺後豎立，使得平面 \(ABD\) 與平面 \(CBD\) **互相垂直**。試求 \(A,C\) 兩點（在空間中）的距離 \(\overline{AC}\)。（最簡根式）",
        "ans": r"**\(\sqrt{337}\) 公分**",
        "key": [r"對角線 \(BD=\sqrt{15^2+20^2}=25\)；點 \(A\) 到 \(BD\) 的垂足距 \(=\tfrac{15\times20}{25}=12\)、垂足離 \(B\) 為 \(\tfrac{15^2}{25}=9\)；同理 \(C\) 的垂足距 \(=12\)、離 \(B\) 為 \(16\)；",
                r"沿 \(BD\) 摺成 **直二面角**，以 \(B\) 為原點、\(BD\) 為 \(x\) 軸建坐標：\(A=(9,12,0)\)、\(C=(16,0,12)\)（兩垂直方向分屬兩平面）；",
                r"\(\overline{AC}=\sqrt{(16-9)^2+12^2+12^2}=\sqrt{49+144+144}=\sqrt{337}\)。"],
        "links": [(r"空間向量 ▸ 考點1 坐標與距離", SPACE, "kp1"), (r"空間向量 ▸ 考點6 二面角與立體", SPACE, "kp6")],
    },
    "114|數B|4": {
        "body": r"某商店推出抽獎活動，提供香蕉、鳳梨、蘋果、橘子 **四種** 不同款式的水果公仔當獎品。每次抽獎可得 \(1\) 個公仔，每種款式被抽中的機率皆相等。某甲決定抽獎 **四次**，試問他恰抽到 **三種不同款式** 公仔的機率為何？",
        "opts": [r"\(\dfrac{5}{16}\)", r"\(\dfrac{3}{8}\)", r"\(\dfrac12\)", r"\(\dfrac{9}{16}\)", r"\(\dfrac58\)"],
        "ans": r"**(4)** \(\dfrac{9}{16}\)",
        "key": [r"四次抽獎、每次 \(4\) 種等機率 → 樣本空間 \(4^4=256\)；",
                r"恰三款 → 款式分布為 \((2,1,1)\)：選 \(3\) 款 \(C^4_3=4\)、選誰重複 \(3\) 種、排列 \(\tfrac{4!}{2!}=12\)，共 \(4\times3\times12=144\)；",
                r"機率 \(=\dfrac{144}{256}=\dfrac{9}{16}\)，選 (4)。"],
        "links": [(r"排列機率 ▸ 考點3 古典機率", PROB, "kp3"), (r"排列機率 ▸ 考點2 組合與分組", PROB, "kp2")],
    },
    "115|數A|3": {
        "body": r"設 \(f(x)=a^x\)（\(a>0\)）。已知 \(c_1,c_2,c_3\) 是公差為 \(\tfrac{10}{3}\) 的等差數列，且 \(f(c_1),f(c_2),f(c_3)\) 是公比為 \(4\) 的等比數列。則等比數列 \(f(10),f(8),f(6)\) 的公比為何？",
        "opts": [r"\(2^{-6/5}\)", r"\(2^{-3/5}\)", r"\(2^{3/5}\)", r"\(2^{6/5}\)", r"\(2^{5/3}\)"],
        "ans": r"**(1)** \(2^{-6/5}\)",
        "key": [r"等差 \(\to\) 等比：公比 \(=a^{\,c_2-c_1}=a^{10/3}=4=2^2\Rightarrow a=2^{3/5}\)；",
                r"\(f(10),f(8),f(6)\) 指數每次差 \(-2\)，公比 \(=a^{-2}=(2^{3/5})^{-2}=2^{-6/5}\)，選 (1)。"],
        "links": [(r"指數對數 ▸ 考點1 指數律", EXPLOG, "kp1"), (r"數列 ▸ 考點2 等比數列", SEQ, "kp2")],
    },
    "110|數A|2": {
        "body": r"五項實數數列 \(a_1,a_2,a_3,a_4,a_5\) 的每一項都大於 \(1\)，且每相鄰兩項中都有一數是另一數的 **兩倍**。若 \(a_1=\log_{10}36\)，則 \(a_5\) 有多少種可能的值？",
        "opts": [r"\(3\)", r"\(4\)", r"\(5\)", r"\(7\)", r"\(8\)"],
        "ans": r"**(1)** \(3\) 種",
        "key": [r"\(a_1=\log 36\approx1.56\)（介於 \(1,2\) 之間）；設各項 \(a_n=a_1\cdot2^{\,e_n}\)，\(e_1=0\)、相鄰 \(e\) 差 \(1\)；",
                r"各項 \(>1\Rightarrow2^{e_n}>\tfrac{1}{\log36}\approx0.64\Rightarrow e_n\ge0\)（不能再減半到 \(<1\)）；",
                r"\(e\) 走 \(4\) 步且不低於 \(0\)，故 \(e_5\in\{0,2,4\}\) → \(a_5\) 有 \(3\) 種，選 (1)。"],
        "links": [(r"數列 ▸ 考點2 等比·倍增", SEQ, "kp2"), (r"指數對數 ▸ 考點2 對數運算", EXPLOG, "kp2")],
    },
    "107|數A|5": {
        "body": r"試問共有幾個角度 \(\theta\) 滿足 \(0^\circ<\theta<180^\circ\)，且 \(\cos(3\theta-60^\circ),\ \cos3\theta,\ \cos(3\theta+60^\circ)\) 依序成 **等差數列**？",
        "opts": [r"\(1\) 個", r"\(2\) 個", r"\(3\) 個", r"\(4\) 個", r"\(5\) 個"],
        "ans": r"**(3)** \(3\) 個",
        "key": [r"等差 \(\Rightarrow2\cos3\theta=\cos(3\theta-60^\circ)+\cos(3\theta+60^\circ)\)；",
                r"和差化積：右式 \(=2\cos3\theta\cos60^\circ=\cos3\theta\)，故 \(2\cos3\theta=\cos3\theta\Rightarrow\cos3\theta=0\)；",
                r"\(3\theta=90^\circ,270^\circ,450^\circ\Rightarrow\theta=30^\circ,90^\circ,150^\circ\)，共 \(3\) 個，選 (3)。"],
        "links": [(r"三角 ▸ 考點5 和差倍角", TRIG, "kp5"), (r"數列 ▸ 考點1 等差數列", SEQ, "kp1")],
    },
    "106|數A|A": {
        "body": r"遞迴數列 \(\langle a_n\rangle\) 滿足 \(a_n=a_{n-1}+f(n-2)\)（\(n\ge2\)），其中 \(f(x)\) 為 **二次多項式**。若 \(a_1=1,\ a_2=2,\ a_3=5,\ a_4=12\)，則 \(a_5=\) ？",
        "ans": r"**\(25\)**",
        "key": [r"由相鄰差求 \(f\)：\(f(0)=a_2-a_1=1\)、\(f(1)=a_3-a_2=3\)、\(f(2)=a_4-a_3=7\)；",
                r"設 \(f(x)=ax^2+bx+c\)：\(c=1\)、\(a+b=2\)、\(2a+b=3\Rightarrow a=1,b=1\)，故 \(f(x)=x^2+x+1\)；",
                r"\(a_5=a_4+f(3)=12+(9+3+1)=25\)。"],
        "links": [(r"數列 ▸ 考點3 遞迴數列", SEQ, "kp3"), (r"多項式 ▸ 考點2 二次函數", POLY, "kp2")],
    },
    "113|數A|6": {
        "body": r"在同一平面上，相距 \(7\) 公里的 \(A,B\) 兩砲台，\(A\) 在 \(B\) 的正東方。某次演習 \(A\) 向西偏北 \(\theta\)、\(B\) 向東偏北 \(\theta\)（\(\theta\) 為銳角）發射砲彈，皆命中 \(9\) 公里外同一目標 \(P\)。接著 \(A\) 改向西偏北 \(\tfrac{\theta}{2}\) 發射，彈著點為 \(9\) 公里外的 \(Q\)。試問 \(\overline{BQ}\) 為何？",
        "fig": r'''<svg viewBox="0 0 172 152" width="172" height="152" xmlns="http://www.w3.org/2000/svg" font-family="system-ui,'Segoe UI',sans-serif">
  <line x1="35" y1="135" x2="112" y2="135" stroke="#9aa6b2" stroke-width="1.4"/>
  <line x1="35" y1="135" x2="73" y2="44" stroke="#2c7a7b" stroke-width="1.5"/>
  <line x1="112" y1="135" x2="73" y2="44" stroke="#2c7a7b" stroke-width="1.5"/>
  <line x1="112" y1="135" x2="29" y2="80" stroke="#2c7a7b" stroke-width="1.4" stroke-dasharray="4 3"/>
  <line x1="35" y1="135" x2="29" y2="80" stroke="#c0392b" stroke-width="2.5"/>
  <g fill="#2c3e50"><circle cx="35" cy="135" r="3"/><circle cx="112" cy="135" r="3"/><circle cx="73" cy="44" r="3"/><circle cx="29" cy="80" r="3"/></g>
  <g font-size="12" font-weight="bold" font-style="italic" fill="#22404a"><text x="23" y="143">B</text><text x="115" y="143">A</text><text x="70" y="38">P</text><text x="12" y="83">Q</text></g>
  <g font-size="10" fill="#2c7a7b"><text x="46" y="88">9</text><text x="96" y="88">9</text></g>
  <text x="66" y="148" font-size="10" fill="#777">7</text>
  <text x="30" y="108" font-size="10" fill="#c0392b" font-weight="bold">BQ?</text>
</svg>''',
        "opts": [r"\(4\) 公里", r"\(4.5\) 公里", r"\(5\) 公里", r"\(5.5\) 公里", r"\(6\) 公里"],
        "ans": r"**(3)** \(5\) 公里",
        "key": [r"\(P\) 距 \(A,B\) 皆 \(9\)、\(AB=7\) → \(P\) 在 \(AB\) 中垂線上，\(A\) 到中點 \(3.5\)，故 \(\cos\theta=\dfrac{3.5}{9}=\dfrac{7}{18}\)；",
                r"半角：\(\cos\dfrac\theta2=\sqrt{\dfrac{1+\cos\theta}{2}}=\dfrac56\)、\(\sin\dfrac\theta2=\dfrac{\sqrt{11}}6\)；",
                r"設 \(B=(0,0),A=(7,0)\)：\(Q=A+9(-\cos\tfrac\theta2,\sin\tfrac\theta2)=(-\tfrac12,\tfrac{3\sqrt{11}}2)\)，\(\overline{BQ}=\sqrt{\tfrac14+\tfrac{99}4}=\sqrt{25}=5\)，選 (3)。"],
        "links": [(r"三角 ▸ 考點3 三角測量", TRIG, "kp3"), (r"三角 ▸ 考點5 和差·半角", TRIG, "kp5")],
    },
    "115|數A|17": {
        "body": r"直角三角形 \(ABC\) 中 \(\angle CAB=90^\circ\)，\(\overline{AB}\) 上一點 \(D\) 滿足 \(\angle BCD=2\angle ACD\)，且 \(\overline{BC}=2\,\overline{DB}\)。若 \(\overrightarrow{AD}=k\,\overrightarrow{AB}\)，則 \(k=\) ？（最簡分數）",
        "fig": r'''<svg viewBox="0 0 186 150" width="186" height="150" xmlns="http://www.w3.org/2000/svg" font-family="system-ui,'Segoe UI',sans-serif">
  <polygon points="28,122 162,122 28,28" fill="#eef6f4" stroke="#1f6f78" stroke-width="1.6"/>
  <line x1="28" y1="28" x2="64" y2="122" stroke="#c0392b" stroke-width="1.5"/>
  <path d="M28,112 L38,112 L38,122" fill="none" stroke="#1f6f78" stroke-width="1.2"/>
  <g fill="#2c3e50"><circle cx="28" cy="122" r="2.6"/><circle cx="162" cy="122" r="2.6"/><circle cx="28" cy="28" r="2.6"/><circle cx="64" cy="122" r="2.6"/></g>
  <g font-size="13" font-weight="bold" font-style="italic" fill="#22404a"><text x="15" y="134">A</text><text x="166" y="134">B</text><text x="14" y="31">C</text><text x="58" y="137">D</text></g>
  <text x="40" y="66" font-size="9.5" fill="#c0392b">∠BCD=2∠ACD</text>
</svg>''',
        "ans": r"**\(k=\dfrac{3}{11}\)**",
        "key": [r"令 \(\angle ACD=\varphi\)，則 \(\angle ACB=3\varphi\)。直角 \(\triangle\) 中 \(AD=AC\tan\varphi,\ AB=AC\tan3\varphi,\ BC=\dfrac{AC}{\cos3\varphi}\)；",
                r"條件 \(BC=2DB=2(AB-AD)\)：\(\dfrac1{\cos3\varphi}=2(\tan3\varphi-\tan\varphi)=\dfrac{2\sin2\varphi}{\cos3\varphi\cos\varphi}\Rightarrow1=4\sin\varphi\Rightarrow\sin\varphi=\dfrac14\)；",
                r"\(k=\dfrac{AD}{AB}=\dfrac{\tan\varphi}{\tan3\varphi}\)；由 \(\sin\varphi=\tfrac14\) 得 \(\sin3\varphi=\tfrac{11}{16},\cos3\varphi=\tfrac{3\sqrt{15}}{16}\)，算得 \(k=\dfrac{3}{11}\)。"],
        "links": [(r"三角 ▸ 考點5 倍角·和差", TRIG, "kp5"), (r"三角 ▸ 考點2 正餘弦定理", TRIG, "kp2")],
    },
    "109|數A|G": {
        "body": r"設計師以不銹鋼片製成月形：圓弧 \(QRT\) 是以 \(O\) 為圓心、\(QT\) 為直徑的半圓，\(\overline{QT}=2\sqrt3\)；圓弧 \(QST\) 的圓心為 \(P\)，\(\overline{PQ}=\overline{PT}=2\)。兩弧所圍灰色區域 \(QRTSQ\) 的面積為 \(a\pi+\sqrt{b}\)（\(a\) 為有理數、\(b\) 為整數），求 \(a,b\)。",
        "fig": r'''<svg viewBox="0 0 150 184" width="150" height="184" xmlns="http://www.w3.org/2000/svg" font-family="system-ui,'Segoe UI',sans-serif">
  <path d="M60,38 A52,52 0 0 1 60,142 A60,60 0 0 0 60,38 Z" fill="#c4c9cf" fill-opacity="0.55"/>
  <path d="M60,38 A52,52 0 0 1 60,142" fill="none" stroke="#1f6f78" stroke-width="1.7"/>
  <path d="M60,38 A60,60 0 0 1 60,142" fill="none" stroke="#c0392b" stroke-width="1.7"/>
  <line x1="60" y1="38" x2="60" y2="142" stroke="#9aa6b2" stroke-width="1.1" stroke-dasharray="3 3"/>
  <g fill="#2c3e50"><circle cx="60" cy="38" r="2.4"/><circle cx="60" cy="142" r="2.4"/><circle cx="60" cy="90" r="1.9"/><circle cx="30" cy="90" r="2.4"/><circle cx="90" cy="90" r="2.4"/><circle cx="112" cy="90" r="2.4"/></g>
  <g font-size="11" font-weight="bold" font-style="italic" fill="#22404a"><text x="55" y="32">Q</text><text x="55" y="156">T</text><text x="48" y="94">O</text><text x="19" y="94">P</text><text x="84" y="103">S</text><text x="115" y="94">R</text></g>
</svg>''',
        "ans": r"**\(a=\dfrac16,\ b=3\)**（面積 \(=\dfrac\pi6+\sqrt3\)）",
        "key": [r"\(\triangle PQT\)：\(PQ=PT=2,\ QT=2\sqrt3\)，餘弦定理 \(\cos\angle QPT=\dfrac{4+4-12}{8}=-\dfrac12\Rightarrow\angle QPT=120^\circ\)；\(OP=\sqrt{2^2-(\sqrt3)^2}=1\)；",
                r"半圓 \(QRT=\dfrac12\pi(\sqrt3)^2=\dfrac{3\pi}2\)；扇形 \(PQST=\dfrac13\pi\cdot2^2=\dfrac{4\pi}3\)；\(\triangle PQT=\dfrac12\cdot2\sqrt3\cdot1=\sqrt3\)；",
                r"灰色 \(=\) 半圓 \(+\triangle PQT-\) 扇形 \(=\dfrac{3\pi}2+\sqrt3-\dfrac{4\pi}3=\dfrac\pi6+\sqrt3\)，故 \(a=\tfrac16,\ b=3\)。"],
        "links": [(r"三角 ▸ 考點1 弧度·扇形面積", TRIG, "kp1"), (r"三角 ▸ 考點2 餘弦定理", TRIG, "kp2")],
    },
    "111|數A|17": {
        "body": r"坐標空間中一平行六面體，某底面的三頂點為 \((-1,2,1),(-4,3,1),(2,0,-3)\)；另一面有一頂點落在 \(xy\) 平面上且與原點距離為 \(1\)。試求此平行六面體的 **最大體積**。",
        "fig": r'''<svg viewBox="0 0 200 150" width="200" height="150" xmlns="http://www.w3.org/2000/svg" font-family="system-ui,'Segoe UI',sans-serif">
  <path d="M78,96 L163,96 M78,96 L78,50 M78,96 L45,118" fill="none" stroke="#c9d0d6" stroke-width="1" stroke-dasharray="3 3"/>
  <polygon points="45,72 130,72 163,50 78,50" fill="#dcefe9" stroke="#1f6f78" stroke-width="1.4"/>
  <polygon points="130,118 163,96 163,50 130,72" fill="#cfe8e6" stroke="#1f6f78" stroke-width="1.4"/>
  <polygon points="45,118 130,118 130,72 45,72" fill="#eef6f4" stroke="#1f6f78" stroke-width="1.6"/>
  <g fill="#c0392b"><circle cx="45" cy="118" r="2.6"/><circle cx="130" cy="118" r="2.6"/><circle cx="45" cy="72" r="2.6"/></g>
  <text x="150" y="128" font-size="10" fill="#777" font-style="italic">平行六面體</text>
</svg>''',
        "ans": r"**\(21\)**",
        "key": [r"底面平行四邊形面積 \(=|\overrightarrow{AB}\times\overrightarrow{AC}|=\sqrt{181}\)（\(A,B,C\) 為已知三頂點）；",
                r"落在 \(xy\) 平面的頂點設為 \((\sin\theta,\cos\theta,0)\)（距原點 \(1\)）；體積 \(=\) 底面積 \(\times\) 點到底面距離（點到平面距離公式，\(\sqrt{181}\) 約掉）\(=|\,8\sin\theta-6\cos\theta-11\,|\)；",
                r"由正餘弦疊合，\(|8\sin\theta-6\cos\theta-11|\) 最大 \(=11+\sqrt{8^2+6^2}=11+10=21\)，故最大體積 \(=21\)。"],
        "links": [(r"空間向量 ▸ 考點3 外積·體積", SPACE, "kp3"), (r"空間向量 ▸ 考點4 點到平面距離", SPACE, "kp4")],
    },
    "113|數A|8": {
        "body": r"對任一正整數 \(n\ge2\)，令 \(T_n\) 表示三邊長為 \(n,\ n+1,\ n+2\) 的三角形。試選出正確的選項。（海龍公式：\(s=\tfrac{a+b+c}2\)，面積 \(=\sqrt{s(s-a)(s-b)(s-c)}\)）",
        "opts": [r"\(T_2\) 為直角三角形", r"\(T_2,T_3,\dots,T_n\) 的周長形成等差數列",
                 r"\(T_n\) 的面積隨 \(n\) 增大而增大", r"每個 \(T_n\) 都是銳角三角形",
                 r"\(T_n\) 的最大內角大於 \(T_{n+1}\) 的最大內角"],
        "ans": r"**(2)(3)(5)**",
        "key": [r"(1) \(T_2=(2,3,4)\)：最大角對 \(4\)，\(\cos=\dfrac{4+9-16}{12}=-\dfrac14<0\) → **鈍角**（非直角），錯；",
                r"(2) 周長 \(=3n+3=3(n+1)\)、公差 \(3\) → 等差 ✓；(3) 海龍面積隨 \(n\) 遞增 ✓；",
                r"(4)(5) 最大角 \(\cos=\dfrac{n^2+(n+1)^2-(n+2)^2}{2n(n+1)}=\dfrac{n-3}{2n}\) 隨 \(n\) **遞增** → 最大角 **遞減**（(5) 對）；\(T_2\) 鈍角、\(T_3\) 直角，非全銳角 → (4) 錯。答 (2)(3)(5)。"],
        "links": [(r"數列 ▸ 考點1 等差數列", SEQ, "kp1"), (r"三角 ▸ 考點2 餘弦定理·海龍", TRIG, "kp2")],
    },
    "113|數B|17": {
        "body": r"在一圓的圓周上取 \(12\) 個等分點，依順時針標 \(1\) 至 \(12\) 號。由這 \(12\) 個點任取 \(3\) 點為頂點所成的三角形中，**三個內角由小到大成等差數列** 的三角形有幾個？",
        "ans": r"**\(76\)**",
        "key": [r"圓內接三角形三內角 \(=15a^\circ,15b^\circ,15c^\circ\)（\(a,b,c\) 為三段弧的等分數，\(a+b+c=12\)）；成等差 \(\Leftrightarrow a,b,c\) 成等差 \(\Leftrightarrow\) 中項 \(=4\)；",
                r"弧組只有 \(\{4,4,4\},\{3,4,5\},\{2,4,6\},\{1,4,7\}\)；",
                r"\(\{4,4,4\}\)（正三角形）有 \(\tfrac{12}{3}=4\) 個；其餘三組各 \(12\times2=24\) 個 → 共 \(4+3\times24=76\)。"],
        "links": [(r"數列 ▸ 考點1 等差數列", SEQ, "kp1"), (r"排列機率 ▸ 考點1 計數原理", PROB, "kp1")],
    },
    "113|數B|4": {
        "body": r"坐標平面上有向量 \(\vec v=(-2,3)\) 及兩點 \(A,B\)，其中 \(A,B\) 的 \(x\)、\(y\) 坐標皆介於 \(0\) 與 \(1\) 之間（\(0\le x,y\le1\)）。試問 \(|\vec v+\overrightarrow{AB}|\) 的 **最大值**為何？",
        "opts": [r"\(\sqrt{13}\)", r"\(\sqrt{17}\)", r"\(3\sqrt2\)", r"\(5\)", r"\(\sqrt2+\sqrt3\)"],
        "ans": r"**(4)** \(5\)",
        "key": [r"設 \(A=(x_1,y_1),B=(x_2,y_2)\)，\(\overrightarrow{AB}=(x_2-x_1,\ y_2-y_1)\)，兩分量皆在 \([-1,1]\)；",
                r"\(\vec v+\overrightarrow{AB}=(-2+(x_2-x_1),\ 3+(y_2-y_1))\)；",
                r"取 \(x_2-x_1=-1,\ y_2-y_1=1\) → \((-3,4)\)，\(|\vec v+\overrightarrow{AB}|=\sqrt{9+16}=5\)（最大），選 (4)。"],
        "links": [(r"平面向量 ▸ 考點1 向量運算·長度", PVEC, "kp1"), (r"數與式 ▸ 考點2 絕對值·範圍", NUMEXPR, "kp2")],
    },
    "115|數B|15": {
        "body": r"某校健檢，全體學生有近視者占 \(\tfrac12\)、有駝背者占 \(\tfrac14\)，設 \(p\) 為「同時近視且駝背」的比率。已知「**有駝背的學生中有近視的比率**」介於 \(\tfrac23\) 與 \(\tfrac56\) 之間。試求 \(p\) 的範圍 \(\square<p<\square\)。",
        "ans": r"**\(\dfrac16<p<\dfrac{5}{24}\)**",
        "key": [r"列聯表：近視且駝背 \(=p\)，駝背且不近視 \(=\tfrac14-p\)（需 \(\ge0\)）；",
                r"「駝背中有近視」\(=\dfrac{p}{1/4}=4p\)；",
                r"由條件 \(\dfrac23<4p<\dfrac56\Rightarrow\dfrac16<p<\dfrac{5}{24}\)。"],
        "links": [(r"數據分析 ▸ 考點4 列聯表·判讀", DATA, "kp4"), (r"排列機率 ▸ 考點4 條件機率", PROB, "kp4")],
    },
    "111|數B|14": {
        "body": r"坐標平面上有一圓，圓心為 \(O\)、半徑為 \(7\)。\(A,B\) 為圓上相異兩點且 \(\overline{AB}=8\)。試求 \(\overrightarrow{OA}\cdot\overrightarrow{OB}\)。",
        "fig": r'''<svg viewBox="0 0 162 150" width="162" height="150" xmlns="http://www.w3.org/2000/svg" font-family="system-ui,'Segoe UI',sans-serif">
  <circle cx="80" cy="86" r="54" fill="#eef6f4" stroke="#1f6f78" stroke-width="1.6"/>
  <line x1="48" y1="42" x2="112" y2="42" stroke="#c0392b" stroke-width="2"/>
  <line x1="80" y1="86" x2="48" y2="42" stroke="#2c7a7b" stroke-width="1.5"/>
  <line x1="80" y1="86" x2="112" y2="42" stroke="#2c7a7b" stroke-width="1.5"/>
  <g fill="#2c3e50"><circle cx="80" cy="86" r="2.6"/><circle cx="48" cy="42" r="2.6"/><circle cx="112" cy="42" r="2.6"/></g>
  <g font-size="12" font-weight="bold" font-style="italic" fill="#22404a"><text x="75" y="100">O</text><text x="38" y="40">A</text><text x="116" y="40">B</text></g>
  <g font-size="10" fill="#777"><text x="56" y="68">7</text><text x="98" y="68">7</text><text x="76" y="36">8</text></g>
</svg>''',
        "ans": r"**\(17\)**",
        "key": [r"\(\overline{AB}^2=|\overrightarrow{OB}-\overrightarrow{OA}|^2=|\overrightarrow{OA}|^2+|\overrightarrow{OB}|^2-2\,\overrightarrow{OA}\cdot\overrightarrow{OB}\)；",
                r"\(8^2=7^2+7^2-2\,\overrightarrow{OA}\cdot\overrightarrow{OB}\Rightarrow64=98-2\,\overrightarrow{OA}\cdot\overrightarrow{OB}\)；",
                r"\(\overrightarrow{OA}\cdot\overrightarrow{OB}=\dfrac{98-64}{2}=17\)。"],
        "links": [(r"平面向量 ▸ 考點3 內積", PVEC, "kp3"), (r"三角 ▸ 考點2 餘弦定理", TRIG, "kp2")],
    },
    "111|數B|6": {
        "body": r"假設地球為半徑 \(R\) 的球體、經線皆為通過兩極的大圓。某地 \(P\) 在 **北緯 \(75^\circ\)**，沿其所在經線往南移動弧長 \(\dfrac{7\pi}{12}R\) 到達 \(Q\)。試問 \(Q\) 的緯度為何？",
        "fig": r'''<svg viewBox="0 0 150 170" width="150" height="170" xmlns="http://www.w3.org/2000/svg" font-family="system-ui,'Segoe UI',sans-serif">
  <circle cx="75" cy="90" r="54" fill="#eef6f4" stroke="#1f6f78" stroke-width="1.5"/>
  <ellipse cx="75" cy="90" rx="54" ry="16" fill="none" stroke="#9aa6b2" stroke-width="1.1"/>
  <path d="M75,36 A21,54 0 0 0 75,144" fill="none" stroke="#c0392b" stroke-width="1.5"/>
  <g fill="#c0392b"><circle cx="62" cy="48" r="3"/><circle cx="65" cy="120" r="3"/></g>
  <g font-size="10" font-weight="bold"><text x="71" y="30" fill="#555">N</text><text x="71" y="160" fill="#555">S</text></g>
  <g font-size="11" font-weight="bold" font-style="italic" fill="#c0392b"><text x="46" y="50">P</text><text x="48" y="124">Q</text></g>
  <text x="104" y="94" font-size="8.5" fill="#9aa6b2">赤道</text>
</svg>''',
        "opts": [r"北緯 \(75^\circ\)", r"北緯 \(30^\circ\)", r"南緯 \(30^\circ\)", r"南緯 \(45^\circ\)", r"南緯 \(75^\circ\)"],
        "ans": r"**(3)** 南緯 \(30^\circ\)",
        "key": [r"弧長 \(=R\theta\Rightarrow\theta=\dfrac{7\pi}{12}=105^\circ\)（大圓圓心角）；",
                r"從 \(P\)（北緯 \(75^\circ\)）沿經線往南：先 \(75^\circ\) 到赤道，再 \(105^\circ-75^\circ=30^\circ\) 進入南半球；",
                r"故 \(Q\) 在 **南緯 \(30^\circ\)**，選 (3)。"],
        "links": [(r"三角 ▸ 考點1 弧度·弧長", TRIG, "kp1"), (r"空間向量 ▸ 考點1 空間坐標·球面", SPACE, "kp1")],
    },
    "111|數B|4": {
        "body": r"在坐標平面上，已知向量 \(\overrightarrow{PQ}=\left(\log\tfrac15,\ -10^{-5}\right)\)，其中點 \(P\) 的坐標為 \(\left(\log\tfrac12,\ 2^{-5}\right)\)。試問點 \(Q\) 在第幾象限？",
        "opts": [r"第一象限", r"第二象限", r"第三象限", r"第四象限", r"坐標軸上"],
        "ans": r"**(2)** 第二象限",
        "key": [r"設 \(Q=(x,y)\)，由 \(\overrightarrow{PQ}=Q-P\)：\(x-\log\tfrac12=\log\tfrac15\)、\(y-2^{-5}=-10^{-5}\)；",
                r"\(x=\log\tfrac12+\log\tfrac15=\log\tfrac1{10}=-1<0\)（對數律相加）；",
                r"\(y=2^{-5}-10^{-5}=\dfrac1{32}-\dfrac1{100000}>0\)；故 \(x<0,\ y>0\Rightarrow\) **第二象限**，選 (2)。"],
        "links": [(r"平面向量 ▸ 考點1 向量坐標運算", PVEC, "kp1"), (r"指數對數 ▸ 考點2 對數運算", EXPLOG, "kp2")],
    },
    "114|數B|13": {
        "body": r"某景點旁有兩個停車場。假設某日 **任一停車場沒有空位** 的機率為 \(0.7\)，且兩停車場是否有空位 **互不影響**。若一輛車當天來到這兩個停車場外圍，則 **至少有一個停車場內有空位** 的機率為何？（最簡分數）",
        "ans": r"**\(\dfrac{51}{100}\)**",
        "key": [r"「至少一個有空位」的 **餘事件** ＝「兩個都客滿」；",
                r"兩場 **獨立**、各客滿機率 \(0.7\)：兩個都客滿 \(=0.7\times0.7=0.49\)；",
                r"故 \(P=1-0.49=0.51=\dfrac{51}{100}\)。"],
        "links": [(r"排列機率 ▸ 考點5 獨立事件·餘事件", PROB, "kp5"), (r"排列機率 ▸ 考點3 古典機率", PROB, "kp3")],
    },
}


def _load_rows():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["跨單元清單"]
    rows = {}
    for r in range(2, ws.max_row + 1):
        v = [ws.cell(r, c).value for c in range(1, 9)]
        if not any(x is not None for x in v):
            continue
        yr, juan, num, typ, main, cross, concept, diff = v
        key = f"{yr}|{juan}|{str(num)}"
        rows[key] = dict(yr=yr, juan=juan, num=str(num), typ=typ or "",
                         main=main or "", cross=cross or "",
                         concept=concept or "", diff=(diff or "").strip())
    return rows


def _utag(name):
    f = UNIT_FILE.get(name)
    if f:
        return f'<a class="utag" href="{f}">{name}</a>'
    return f'<span class="utag dead">{name}</span>'


def _card(row):
    d = DIFF_CLASS.get(row["diff"], "mid")
    juan = row["juan"]
    return (
        f'<div class="xq" data-juan="{juan}">'
        f'<div class="xq-head">'
        f'<span class="yr">{row["yr"]} {juan}</span>'
        f'<span class="qn">{row["typ"]} {row["num"]}</span>'
        f'<span class="lv {d}">{row["diff"]}</span>'
        f'</div>'
        f'<div class="xq-units">{_utag(row["main"])}<span class="x">×</span>{_utag(row["cross"])}</div>'
        f'<div class="xq-concept">{row["concept"]}</div>'
        f'</div>'
    )


def _rich_card(row, dt):
    d = DIFF_CLASS.get(row["diff"], "mid")
    juan = row["juan"]
    opts = ""
    if dt.get("opts"):
        opts = ('<div class="opts">'
                + "".join(f"<span>({i}) {html_rich(o)}</span>" for i, o in enumerate(dt["opts"], 1))
                + "</div>")
    fig = f'<div class="xq-fig">{dt["fig"]}</div>' if dt.get("fig") else ""
    keyhtml = "".join(f"<p>{html_rich(k)}</p>" for k in dt["key"])
    links = '<span class="x">×</span>'.join(
        f'<a class="utag" href="{f}#{a}">{lab}</a>' for lab, f, a in dt["links"])
    return (
        f'<div class="xq rich" data-juan="{juan}">'
        f'<div class="xq-head" onclick="toggleXq(this)">'
        f'<span class="yr">{row["yr"]} {juan}</span>'
        f'<span class="qn">{row["typ"]} {row["num"]}</span>'
        f'<span class="pair">{row["main"]}×{row["cross"]}</span>'
        f'<span class="lv {d}">{row["diff"]}</span>'
        f'<span class="caret">＋</span>'
        f'</div>'
        f'<div class="xq-detail">'
        f'<div class="xq-q">{html_rich(dt["body"])}{fig}{opts}</div>'
        f'<button class="sol-btn mini" data-s="簡答" data-h="收合簡答" onclick="ts(this)">簡答</button>'
        f'<div class="sol">{html_rich(dt["ans"])}</div>'
        f'<button class="sol-btn mini" data-s="解題關鍵" data-h="收合解題關鍵" onclick="ts(this)">解題關鍵</button>'
        f'<div class="sol">{keyhtml}</div>'
        f'<div class="xq-klink">🔗 直達考點：{links}</div>'
        f'</div></div>'
    )


def build():
    rows = _load_rows()
    used = set()
    sections = []
    for title, ctx, keys in THEMES:
        # 排序：數A 在前、數B 在後；同卷別內依年分倒序（115 最前）
        keys = sorted(keys, key=lambda k: (0 if rows[k]["juan"] == "數A" else 1, -int(rows[k]["yr"])))
        cards = []
        for k in keys:
            if k not in rows:
                raise SystemExit(f"Excel 找不到題目 key：{k}")
            used.add(k)
            cards.append(_rich_card(rows[k], DETAIL[k]) if k in DETAIL else _card(rows[k]))
        n_rich = sum(1 for k in keys if k in DETAIL)
        tag = f'<small>{len(keys)} 題' + (f'·{n_rich} 題可展開' if n_rich else '') + '</small>'
        sections.append(
            f'<div class="part">{title}{tag}</div>'
            f'<div class="callout ctx">{ctx}</div>'
            f'<div class="xq-grid">{"".join(cards)}</div>'
        )
    # 確認沒有漏掉 Excel 裡的題目
    missing = [k for k in rows if k not in used]
    if missing:
        raise SystemExit(f"有 {len(missing)} 題未編入主題：{missing}")

    nA = sum(1 for r in rows.values() if r["juan"] == "數A")
    nB = len(rows) - nA
    css = _asset("style.css")
    extra = """
  .x-intro{background:#fff7ef;border:1px dashed #e0b9a6;border-radius:12px;padding:12px 18px;margin:14px 0 8px;font-size:14.5px;color:#6b5249}
  .x-intro b{color:var(--maroon)}
  .filterbar{display:flex;gap:8px;align-items:center;margin:14px 0 4px;flex-wrap:wrap}
  .filterbar .fb-label{font-weight:700;color:var(--maroon-d);font-size:14px}
  .filterbar button{background:#fff;color:var(--maroon-d);border:1.5px solid var(--maroon);border-radius:20px;padding:5px 16px;font-size:14px;font-weight:700;cursor:pointer;font-family:inherit;transition:.15s}
  .filterbar button.on{background:var(--maroon);color:#fff}
  .callout.ctx{font-size:14.5px;line-height:1.8}
  .xq-grid{display:flex;flex-direction:column;gap:10px;margin:10px 0 6px}
  .xq{background:#fff;border-radius:12px;padding:11px 16px;box-shadow:0 2px 8px rgba(80,40,30,.07);border-left:4px solid var(--btn)}
  .xq[data-juan="數B"]{border-left-color:#b0853a;background:#fffdf7}
  .xq-head{display:flex;align-items:center;gap:9px;flex-wrap:wrap}
  .xq-head .yr{font-weight:800;color:var(--maroon);font-size:15px}
  .xq-head .qn{font-size:12.5px;color:#6b5249;background:#f3e3e8;border-radius:6px;padding:1px 8px}
  .xq-head .lv{margin-left:auto;font-size:12px;font-weight:700;border-radius:6px;padding:1px 9px;color:#fff}
  .lv.easy{background:#3a9d6b}.lv.mid{background:#d39a2a}.lv.hard{background:#c0392b}
  .xq-units{margin:5px 0 4px;font-size:15px;font-weight:700}
  .xq-units .x{color:#b08;margin:0 6px;font-weight:800}
  .utag{color:var(--btn);text-decoration:none;border-bottom:1.5px solid #bfe0dd}
  .utag:hover{background:var(--btn-bg)}
  .utag.dead{color:#999;border:none;font-weight:600}
  .xq-concept{font-size:13.5px;color:#5a4a52;line-height:1.7}
  .xstat{display:flex;gap:18px;flex-wrap:wrap;font-size:13.5px;color:#6b5249;margin:6px 0 2px}
  .xstat b{color:var(--maroon);font-size:18px}
  /* 可展開的互動題卡 */
  .xq.rich .xq-head{cursor:pointer;user-select:none}
  .xq.rich:hover{box-shadow:0 3px 12px rgba(31,111,120,.16)}
  .xq.rich .pair{font-size:12.5px;color:#1f6f78;font-weight:700;background:#e9f5f4;border-radius:6px;padding:1px 9px}
  .xq.rich .caret{font-weight:800;color:#fff;background:var(--btn);width:20px;height:20px;border-radius:50%;display:inline-flex;align-items:center;justify-content:center;font-size:15px;line-height:1}
  .xq-detail{display:none;margin-top:11px;border-top:1px dashed var(--line);padding-top:11px}
  .xq-detail.open{display:block}
  .xq-q{font-size:15.5px;color:#2b2b2b;line-height:1.9;background:#fbf8f5;border-radius:8px;padding:11px 14px}
  .xq-q .opts{margin-top:8px}
  .xq-q .opts>span{display:block;font-size:14.5px;color:#444;padding:2px 0}
  .xq-fig{text-align:center;margin:10px 0 2px}
  .xq-fig svg{max-width:240px}
  .xq .sol-btn{margin:8px 8px 2px 0}
  .xq .sol{font-size:14.5px}
  .xq .sol p{margin:.35em 0}
  .xq-klink{margin-top:11px;font-size:13px;color:#6b5249;line-height:2}
  .xq-klink .utag{display:inline-block;margin:2px 0;border-bottom-color:#bfe0dd}
"""
    body = f"""<div class="hero"><h1>跨單元整合 · 脈絡地圖</h1>
<p>學測數學 106–115　共 {len(rows)} 道「跨單元」考題，依<b>結合的概念主軸</b>分成五大脈絡</p></div>
<div class="x-intro">
<b>為什麼要看跨單元題？</b>　近年學測越來越愛把兩個單元縫在一起——會單一公式不夠，要看出「這題其實是 A × B」。
這頁把 106–115 的跨單元題依<b>「誰跟誰結合、為什麼結合」</b>排成五條脈絡。標有 <span style="color:#fff;background:#1f6f78;border-radius:50%;padding:0 6px;font-weight:800">＋</span> 的數A題卡<b>點一下可展開</b>完整題目，並有<b>簡答、解題關鍵</b>按鈕與<b>直達考點</b>連結。（向量脈絡已完成，其餘陸續補上。）
<div class="xstat"><span>總題數 <b>{len(rows)}</b></span><span>數A <b>{nA}</b></span><span>數B <b>{nB}</b></span><span>五大脈絡主軸</span></div>
</div>
<div class="filterbar"><span class="fb-label">篩選卷別：</span>
<button data-f="all" class="on" onclick="flt(this,'all')">全部</button>
<button data-f="數A" onclick="flt(this,'數A')">只看數A</button>
<button data-f="數B" onclick="flt(this,'數B')">只看數B</button></div>
{''.join(sections)}
<div class="foot">資料整理自《學測數學考題分析 106–115》；分組與脈絡敘述為教學編排，題幹與官方答案以原卷為準。</div>"""

    # 上方導覽：回首頁 + 四大主題單元下拉
    try:
        from units import SECTIONS, UNITS
        by = {u["slug"]: u for u in UNITS}
        opts = ['<option value="">單元 ▾</option>', '<option value="index.html">📚 目錄首頁</option>',
                '<option value="115學測數學_概念地圖.html">🗺️ 概念地圖</option>']
        for label, slugs in SECTIONS:
            grp = "".join(f'<option value="{by[s]["file"]}">{by[s]["emoji"]} {by[s]["title"]}</option>'
                          for s in slugs if s in by and not by[s].get("draft"))
            if grp:
                opts.append(f'<optgroup label="{label}">{grp}</optgroup>')
        navsel = ('<select class="nav-select" onchange="if(this.value){location.href=this.value;}">'
                  + "".join(opts) + '</select>')
    except Exception:
        navsel = '<a class="nav-select" href="index.html">📚 目錄首頁</a>'

    toolbar = (f'<div class="toolbar"><span class="t-title">跨單元整合 · 脈絡地圖</span>{navsel}</div>')
    js = ("function flt(btn,f){document.querySelectorAll('.filterbar button').forEach(b=>b.classList.remove('on'));"
          "btn.classList.add('on');"
          "document.querySelectorAll('.xq').forEach(c=>{c.style.display=(f==='all'||c.dataset.juan===f)?'':'none';});"
          "document.querySelectorAll('.part').forEach(p=>{let g=p.nextElementSibling;"
          "while(g&&!g.classList.contains('xq-grid'))g=g.nextElementSibling;"
          "if(g){let vis=[...g.children].some(c=>c.style.display!=='none');"
          "p.style.display=vis?'':'none';p.nextElementSibling.style.display=vis?'':'none';}});}"
          "function toggleXq(h){var d=h.nextElementSibling;var o=d.classList.toggle('open');"
          "var c=h.querySelector('.caret');if(c)c.textContent=o?'\\u2212':'\\uFF0B';}"
          "function ts(b){var s=b.nextElementSibling;var o=s.classList.toggle('open');"
          "b.classList.toggle('active',o);b.textContent=o?b.dataset.h:b.dataset.s;}")

    html = f"""<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
{ANALYTICS}<title>學測數學 · 跨單元整合脈絡地圖（106–115）</title>
<link rel="stylesheet" href="{KATEX}/katex.min.css">
<style>
{css}
{extra}</style>
</head>
<body>
{toolbar}
<div class="wrap">
{body}
</div>
<script src="{KATEX}/katex.min.js"></script>
<script src="{KATEX}/contrib/auto-render.min.js"></script>
<script>
renderMathInElement(document.body,{{delimiters:[{{left:'\\\\(',right:'\\\\)',display:false}},{{left:'\\\\[',right:'\\\\]',display:true}}]}});
{js}
</script>
</body>
</html>
"""
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    io.open(OUT, "w", encoding="utf-8").write(html)
    print(f"[cross] -> {os.path.basename(OUT)}  （{len(rows)} 題 / 數A {nA} · 數B {nB} / 5 主軸）")


if __name__ == "__main__":
    build()
